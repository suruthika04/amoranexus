import os
import csv
import io
from datetime import datetime
from flask import Flask, request, jsonify, send_file, Response
from flask_cors import CORS
from pymongo import MongoClient
from bson import ObjectId
from bson.json_util import dumps
import openpyxl
from dotenv import load_dotenv
import requests


env_path = os.path.join(os.path.dirname(__file__), '.env')
load_dotenv(dotenv_path=env_path)

app = Flask(__name__)
CORS(app)

# ─── MongoDB Connection ───────────────────────────
MONGO_URI = os.getenv("MONGO_URI", "mongodb://localhost:27017/amora_nexus")
ADMIN_KEY = os.getenv("ADMIN_PASSWORD", "amoranexus2024")

try:
    client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=5000)
    client.server_info()
    db = client.Amora_Nexus_Database
    registrations_col = db.N8N_Masterclass_Registration
    print("Checked: Connected to MongoDB")
except Exception as e:
    print(f"Warning: MongoDB connection failed: {e}")
    db = None
    registrations_col = None


def require_admin(f):
    """Decorator to protect admin routes."""
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        key = request.headers.get("X-Admin-Key", "")
        if key != ADMIN_KEY:
            return jsonify({"error": "Unauthorized"}), 401
        return f(*args, **kwargs)
    return decorated


def serialize_doc(doc):
    """Convert MongoDB doc to JSON-serializable dict."""
    doc["_id"] = str(doc["_id"])
    if "createdAt" in doc and isinstance(doc["createdAt"], datetime):
        doc["createdAt"] = doc["createdAt"].isoformat()
    return doc


# ─── Health Check ─────────────────────────────────
@app.route("/api/health", methods=["GET"])
def health():
    return jsonify({"status": "ok", "message": "Amora Nexus API is running"}), 200


# ─── Register (Public) ────────────────────────────
@app.route("/api/register", methods=["POST"])
def register():
    data = request.get_json()
    print("Received Data:", data)

    if not data:
        return jsonify({"message": "No data provided"}), 400

    required = ["fullName", "email", "phone", "college", "city"]

    for field in required:
        if not data.get(field, "").strip():
            return jsonify({"message": f"{field} is required"}), 400

    doc = {
        "fullName": data.get("fullName", "").strip(),
        "email": data.get("email", "").lower().strip(),
        "phone": data.get("phone", "").strip(),
        "college": data.get("college", "").strip(),
        "department": data.get("department", "").strip(),
        "yearOfStudy": data.get("yearOfStudy", "").strip(),
        "city": data.get("city", "").strip(),
        "createdAt": datetime.utcnow(),
    }

    if registrations_col is None:
        return jsonify({
            "message": "Database connection failed. Please try again later."
        }), 500

    try:
        result = registrations_col.insert_one(doc)
        print("Inserted ID:", result.inserted_id)

        # Send data to n8n
        webhook_data = {
            "fullName": doc["fullName"],
            "email": doc["email"],
            "phone": doc["phone"],
            "college": doc["college"],
            "department": doc["department"],
            "yearOfStudy": doc["yearOfStudy"],
            "city": doc["city"],
            "createdAt": doc["createdAt"].isoformat()
        }

        requests.post(
            "https://amoranexus.app.n8n.cloud/webhook/cff7aab5-00f4-4e81-bfcd-ba3b5420ca28",
            json=webhook_data
        )

        return jsonify({
            "message": "Registration successful!"
        }), 201

    except Exception as e:
        print("Error:", str(e))
        return jsonify({
            "message": str(e)
        }), 500

# ─── Get All Registrations (Admin) ────────────────
@app.route("/api/registrations", methods=["GET"])
@require_admin
def get_registrations():
    if registrations_col is None:
        return jsonify({"registrations": [], "total": 0}), 200

    docs = list(registrations_col.find().sort("createdAt", -1))
    docs = [serialize_doc(d) for d in docs]
    return jsonify({"registrations": docs, "total": len(docs)}), 200


# ─── Delete Registration (Admin) ──────────────────
@app.route("/api/registrations/<id>", methods=["DELETE"])
@require_admin
def delete_registration(id):
    if registrations_col is None:
        return jsonify({"message": "DB not available"}), 503
    try:
        registrations_col.delete_one({"_id": ObjectId(id)})
        return jsonify({"message": "Deleted successfully"}), 200
    except Exception as e:
        return jsonify({"message": str(e)}), 400


# ─── Export Excel (Admin) ─────────────────────────
@app.route("/api/export/excel", methods=["GET"])
@require_admin
def export_excel():
    if registrations_col is None:
        return jsonify({"message": "DB not available"}), 503

    docs = list(registrations_col.find().sort("createdAt", -1))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registrations"

    headers = ["#", "Full Name", "Email", "Phone", "College / Org", "Department", "Year of Study", "City", "Registered At"]
    ws.append(headers)

    # Style header row
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, doc in enumerate(docs, 1):
        ws.append([
            i,
            doc.get("fullName", ""),
            doc.get("email", ""),
            doc.get("phone", ""),
            doc.get("college", ""),
            doc.get("department", ""),
            doc.get("yearOfStudy", ""),
            doc.get("city", ""),
            doc.get("createdAt", "").strftime("%Y-%m-%d %H:%M") if isinstance(doc.get("createdAt"), datetime) else str(doc.get("createdAt", "")),
        ])

    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"amora_nexus_registrations_{datetime.utcnow().strftime('%Y%m%d')}.xlsx",
    )


# ─── Export CSV (Admin) ───────────────────────────
@app.route("/api/export/csv", methods=["GET"])
@require_admin
def export_csv():
    if registrations_col is None:
        return jsonify({"message": "DB not available"}), 503

    docs = list(registrations_col.find().sort("createdAt", -1))

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["#", "Full Name", "Email", "Phone", "College / Org", "Department", "Year of Study", "City", "Registered At"])

    for i, doc in enumerate(docs, 1):
        writer.writerow([
            i,
            doc.get("fullName", ""),
            doc.get("email", ""),
            doc.get("phone", ""),
            doc.get("college", ""),
            doc.get("department", ""),
            doc.get("yearOfStudy", ""),
            doc.get("city", ""),
            doc.get("createdAt", "").strftime("%Y-%m-%d %H:%M") if isinstance(doc.get("createdAt"), datetime) else str(doc.get("createdAt", "")),
        ])

    csv_data = output.getvalue()
    return Response(
        csv_data,
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=amora_nexus_registrations_{datetime.utcnow().strftime('%Y%m%d')}.csv"},
    )

if __name__ == "__main__":
    port = int(os.getenv("PORT", 5000))
    print(f"Amora Nexus API running on http://localhost:{port}")
    app.run(debug=True, host="0.0.0.0", port=port)
